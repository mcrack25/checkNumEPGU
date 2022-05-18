import os
from openpyxl import load_workbook
from packages.dirs import Dirs
from packages.config import Config
from app.models.WhatTable import WhatTable
from app.controllers.Uploader import Uploader


class WhatUploader(Uploader):
    whatDir = ''
    file_path = None
    max_row = 0

    def __init__(self):
        self.whatDir = Dirs().get('what')

    def run(self):
        first_row_what = Config().get('row_what')

        for what_file_one in os.listdir(self.whatDir):
            file_what_path = os.path.join(self.whatDir, what_file_one)
            wb = load_workbook(filename=file_what_path)
            ws = wb.active
            max_row = ws.max_row
            for i in range(first_row_what, max_row + 1):
                cell_coord_A = "A" + str(i)
                cell_coord_B = "B" + str(i)
                cell_coord_C = "C" + str(i)

                snils = ws[cell_coord_A].value
                request_date = ws[cell_coord_B].value
                serviceid = ws[cell_coord_C].value

                file_name = what_file_one
                num_row = i

                self.upload(snils, request_date, serviceid, file_name, num_row)

    def upload(self, snils, request_date, serviceid, file_name, num_row):
        snils = self.validString(snils)
        request_date = self.validString(request_date)
        serviceid = self.validString(serviceid)

        result = (WhatTable.insert(
            snils=snils,
            request_date=request_date,
            serviceid=serviceid,
            file_name=file_name,
            num_row=num_row
        ).on_conflict('replace').execute())