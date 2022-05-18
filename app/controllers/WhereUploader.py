import os
from openpyxl import load_workbook
from packages.dirs import Dirs
from packages.config import Config
from app.models.WhereTable import WhereTable
from app.controllers.Uploader import Uploader

class WhereUploader(Uploader):
    whereDir = ''
    file_path = None
    max_row = 0

    def __init__(self):
        self.whereDir = Dirs().get('where')

    def run(self):
        first_row_where = Config().get('row_where')

        for where_file_one in os.listdir(self.whereDir):
            file_where_path = os.path.join(self.whereDir, where_file_one)
            wb = load_workbook(filename=file_where_path)
            ws = wb.active
            max_row = ws.max_row
            for i in range(first_row_where, max_row + 1):
                cell_coord = "B" + str(i)

                snils = ws[cell_coord].value
                file_name = where_file_one
                num_row = i

                self.upload(snils, file_name, num_row)

    def upload(self, snils, file_name, num_row):
        snils = self.validString(snils)

        result = (WhereTable.insert(
            snils=snils,
            file_name=file_name,
            num_row=num_row
        ).on_conflict('replace').execute())