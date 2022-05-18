import os
import openpyxl
from packages.config import Config
from packages.dirs import Dirs

class ResultFile:
    file_path = None
    wb = None
    ws = None
    num_row = 1
    num_rec = 1
    config = {}
    rows = []

    def __init__(self):
        self.config = Config().getAll()
        resultDir = Dirs().get('result')
        self.file_path = os.path.join(resultDir, self.config['file_name'])
        self.makeExcel()
        self.setWidth()

    def makeExcel(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.config['list_name']
        self.wb = wb
        self.ws = ws

    def saveFile(self):
        for row in self.rows:
            self.contentRowPutInsert(row)
        self.wb.save(self.file_path)

    def setWidth(self):
        ws = self.ws
        sizes = self.config['sizes']

        rdA = ws.column_dimensions['A']
        rdA.width = sizes['n1']

        rdB = ws.column_dimensions['B']
        rdB.width = sizes['n2']

        rdC = ws.column_dimensions['C']
        rdC.width = sizes['n3']

        rdD = ws.column_dimensions['D']
        rdD.width = sizes['n4']

        rdE = ws.column_dimensions['E']
        rdE.width = sizes['n5']

        rdF = ws.column_dimensions['F']
        rdF.width = sizes['n6']

        rdG = ws.column_dimensions['G']
        rdG.width = sizes['n7']

        rdH = ws.column_dimensions['H']
        rdH.width = sizes['n8']

        self.ws = ws

    def headerContentPut(self):
        ws = self.ws
        num_row = self.num_row
        headers = self.config['headers']

        cell_coord_A = "A" + str(num_row)
        a = ws[cell_coord_A]
        a.value = headers['n1']
        a.font = a.font.copy(bold=True)
        a.alignment = a.alignment.copy(horizontal='center', wrapText=True)

        cell_coord_B = "B" + str(num_row)
        b = ws[cell_coord_B]
        b.value = headers['n2']
        b.font = b.font.copy(bold=True)
        b.alignment = b.alignment.copy(horizontal='center', wrapText=True)

        cell_coord_C = "C" + str(num_row)
        c = ws[cell_coord_C]
        c.value = headers['n3']
        c.font = c.font.copy(bold=True)
        c.alignment = c.alignment.copy(horizontal='center', wrapText=True)

        cell_coord_D = "D" + str(num_row)
        d = ws[cell_coord_D]
        d.value = headers['n4']
        d.font = d.font.copy(bold=True)
        d.alignment = d.alignment.copy(horizontal='center', wrapText=True)

        cell_coord_E = "E" + str(num_row)
        e = ws[cell_coord_E]
        e.value = headers['n5']
        e.font = e.font.copy(bold=True)
        e.alignment = e.alignment.copy(horizontal='center', wrapText=True)

        cell_coord_F = "F" + str(num_row)
        f = ws[cell_coord_F]
        f.value = headers['n6']
        f.font = f.font.copy(bold=True)
        f.alignment = f.alignment.copy(horizontal='center', wrapText=True)

        cell_coord_G = "G" + str(num_row)
        g = ws[cell_coord_G]
        g.value = headers['n7']
        g.font = g.font.copy(bold=True)
        g.alignment = g.alignment.copy(horizontal='center', wrapText=True)

        cell_coord_H = "H" + str(num_row)
        h = ws[cell_coord_H]
        h.value = headers['n8']
        h.font = h.font.copy(bold=True)
        h.alignment = h.alignment.copy(horizontal='center', wrapText=True)

        self.num_row += 1

    def contentRowPut(self, row):
        self.rows.append(row)

    def contentRowPutInsert(self, row):
        ws = self.ws
        num_row = self.num_row
        num_rec = self.num_rec

        cell_coord_A = "A" + str(num_row)
        a = ws[cell_coord_A]
        a.value = num_rec

        cell_coord_B = "B" + str(num_row)
        b = ws[cell_coord_B]
        b.value = row['what_file']

        cell_coord_C = "C" + str(num_row)
        c = ws[cell_coord_C]
        c.value = row['what_row']

        cell_coord_D = "D" + str(num_row)
        d = ws[cell_coord_D]
        d.value = row['where_file']

        cell_coord_E = "E" + str(num_row)
        e = ws[cell_coord_E]
        e.value = row['where_row']

        cell_coord_F = "F" + str(num_row)
        f = ws[cell_coord_F]
        f.value = row['value']

        cell_coord_G = "G" + str(num_row)
        g = ws[cell_coord_G]
        g.value = row['request_date']

        cell_coord_H = "H" + str(num_row)
        h = ws[cell_coord_H]
        h.value = row['serviceid']

        self.num_row+=1
        self.num_rec+=1
